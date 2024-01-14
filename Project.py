import openpyxl
import xlsxwriter

def write_to_excel(filename, sheet_name, data):
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet(sheet_name)
    for row_num, row_data in enumerate(data):
        for col_num, cell_value in enumerate(row_data):
            worksheet.write(row_num, col_num, cell_value)

    workbook.close()
    return True
 
def read_excel(filename, sheet_name):
    try:
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook[sheet_name]
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))

 

        return data
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None
    
filename='231127.xlsx'
sheet_name='A1'
data = read_excel(filename, sheet_name)
data_2 = data



Quota_name ="Q1_Quota"
brand_marker ="brand_"
Question_id = "Q1"
Quota_marker_list = []
final_marker = []

q1_columns = [col for col in data[0] if col.startswith('Q1')]
q1_columns_index = [index for index,col in enumerate(data[0]) if col.startswith('Q1')]
for x in range(1, len(q1_columns)+1):
    Quota_marker_list.append(f",/{Quota_name}/{brand_marker}{x}")


def Create_marker(row, col_index, marker_list  ):
    row_data = []
    row_marker = ''
    for index in range(0, len(col_index)):
        if row[col_index[index]] == 0:
            print("row value 0")
        if row[col_index[index]] == 1:
            row_marker += ''.join(marker_list[index])
    row_data.append(row[0])
    row_data.append(row_marker)

    
    return row_data

final_marker = []
for i in range(1, len(data)):
    print("i", i)
    row = data[i]
    final_marker.append(Create_marker(row, q1_columns_index,Quota_marker_list))
    print(Create_marker(row, q1_columns_index,Quota_marker_list))


filename1='output.xlsx'
sheet_name='A1'

write_to_excel(filename1, sheet_name, final_marker)